from PySide6.QtGui import QDoubleValidator, QIntValidator
from PySide6.QtWidgets import (
    QMainWindow,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QLineEdit,
    QLabel,
    QHBoxLayout,
    QCheckBox,
    QStackedWidget,
    QRadioButton,
    QSpacerItem,
    QSizePolicy, QButtonGroup, QMessageBox, QApplication,
)
from PySide6.QtCore import QRegularExpression
from PySide6.QtGui import QRegularExpressionValidator
from PySide6.QtWidgets import QMessageBox
import random
from PySide6.QtCore import Qt
from MAIN import DashboardWindow

class BasePage(QWidget):
    def __init__(self):
        super().__init__()
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setSpacing(12)
        self.main_layout.setContentsMargins(15, 15, 15, 15)
        self.main_layout.setAlignment(Qt.AlignTop)

        # Navigation layout at bottom
        self.nav_layout = QHBoxLayout()
        self.nav_layout.setContentsMargins(0, 0, 0, 0)

    def add_nav_buttons(self, prev_callback=None, next_callback=None):

        self.prev_btn = QPushButton("Back")
        self.next_btn = QPushButton("Next")
        self.prev_btn.setFixedSize(100,30)
        self.next_btn.setFixedSize(100,30)
        self.prev_btn.setStyleSheet("""
            background-color: #C23731;
            color: #1E1E1E;
            border: 2px solid black;
            border-radius: 15px;  /* bigger = more rounded */
        """)

        self.next_btn.setStyleSheet("""
            background-color: #61AF5E;
            color: #1E1E1E;
            border: 2px solid black;
            border-radius: 15px;  /* bigger = more rounded */
        """)

        if prev_callback:
            self.prev_btn.clicked.connect(prev_callback)
        if next_callback:
            self.next_btn.clicked.connect(next_callback)

        # Add buttons to nav layout with stretch between
        self.nav_layout.addWidget(self.prev_btn, alignment=Qt.AlignLeft)
        self.nav_layout.addStretch()
        self.nav_layout.addWidget(self.next_btn, alignment=Qt.AlignRight)

    def centered_row(self, widget):
        row = QHBoxLayout()
        row.addStretch()
        row.addWidget(widget)
        row.addStretch()
        return row

    def make_field(self, placeholder, width=60, height=20):
        field = QLineEdit()
        field.setPlaceholderText(placeholder)
        field.setFixedSize(width, height)
        field.setStyleSheet("color: black; background: #fff; border-radius: 8px;")
        return field

###########################################################################

class SecondAssignment(QMainWindow):

    def go_prev(self):
        idx = self.stacked.currentIndex()
        if idx > 0:
            self.stacked.setCurrentIndex(idx - 1)

    def go_next(self):
        idx = self.stacked.currentIndex()
        current_page = self.pages[idx]

        # Save inputs
        if isinstance(current_page, Page1):
            self.input_data["page1"]["start"] = int(current_page.start_field.text())
            self.input_data["page1"]["end"] = int(current_page.end_field.text())
            self.input_data["page1"]["probabilities"] = current_page.prob_field.text()
            self.input_data["page1"]["equal"] = current_page.equal_checkbox.isChecked()

        elif isinstance(current_page, Page2):
            self.input_data["page2"]["start"] = int(current_page.start_field.text())
            self.input_data["page2"]["end"] = int(current_page.end_field.text())
            self.input_data["page2"]["probabilities"] = current_page.prob_field.text()
            self.input_data["page2"]["equal"] = current_page.equal_checkbox.isChecked()

        elif isinstance(current_page, Page3):
            from PySide6.QtWidgets import QMessageBox
            text = current_page.num_instances_field.text().strip()
            if not text:
                QMessageBox.warning(self, "Invalid Input", "Please enter the number of instances")
                return
            try:
                self.input_data["page3"]["instances"] = int(text)
            except ValueError:
                QMessageBox.warning(self, "Invalid Input", "Number of instances must be an integer")
                return

            self.input_data["page3"]["traffic_type"] = (
                current_page.traffic_radio.text() if current_page.traffic_radio.isChecked() else "Other"
            )

# --- Page1/Page2 validation ---
        if isinstance(current_page, (Page1, Page2)):
            # Get start/end
            try:
                start = int(current_page.start_field.text())
                end = int(current_page.end_field.text())
            except ValueError:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Invalid Input", "Start and End must be numbers")
                return

            # Get probabilities
            probs_text = current_page.prob_field.text().strip()
            # Split by spaces
            prob_items = [x for x in probs_text.split() if x]  # ignore extra spaces
            try:
                probs = [float(x) for x in prob_items]
            except ValueError:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Invalid Input", "Probabilities must be numbers separated by spaces")
                return

            # Check number of items matches range
            if len(prob_items) != (end - start + 1):
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Invalid Input",
                                    f"Number of probabilities must be equal to end-start+1 ({end - start + 1})")
                return

            # Check sum = 1
            if abs(sum(probs) - 1.0) > 1e-6:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Invalid Input", "Sum of probabilities must equal 1")
                return

            # Save values for later
            current_page.start_val = start
            current_page.end_val = end
            current_page.probs_val = probs

        # If not last page â go next normally
        if idx < self.stacked.count() - 1:
            self.stacked.setCurrentIndex(idx + 1)
        else:
            # On the last page (Page4)
            page4 = self.pages[-1]  # get last page object
            selected = page4.selected_option()

            if not selected:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "No Option Selected", "Please choose an output option first.")
                return

            # Example: handle different selections
            if "Excel" in selected:
                print("Exporting results to Excel...")

            elif "GUI" in selected:
                print("Displaying results in GUI...")
                # show another window
            elif "Terminal" in selected:
                print("Printing results to terminal...")
                p1 = self.input_data["page1"]
                p2 = self.input_data["page2"]
                p3 = self.input_data["page3"]

                probs1 = list(map(float, p1["probabilities"].split()))
                probs2 = list(map(float, p2["probabilities"].split()))

                Temp1 = Page4.generate_interarrival_distribution(p1["start"], p1["end"], probs1)
                Temp2 = Page4.generate_service_time_distribution(p2["start"], p2["end"], probs2)
                queue = Page4.simulate_queue(Temp1, Temp2, p3["instances"])[0]
                Page4.print_table_terminal(
                Temp1, ["Interarrival Time", "Probability", "Cumulative Probability", "Range"]
                )
                Page4.print_table_terminal(
                Temp2, ["Service Time", "Probability", "Cumulative Probability", "Range"]
                )
                Page4.print_table_terminal(
                    queue,
                    ["user", "interarrival_time", "arrival_time", "service_time",
                     "service_begin", "waiting_time", "service_end", "time_in_system", "idle_time"]
                )
                # print(queue[1])
            # Optionally close the window after action
            self.close()

    def mousePressEvent(self, event):
        self._drag_pos = event.globalPosition().toPoint() - self.pos()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)

    def center_on_screen(self):
        # Get screen geometry
        screen = QApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        # Get window geometry
        window_geometry = self.frameGeometry()
        # Calculate center
        center_point = screen_geometry.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())

    def __init__(self, num_pages=3):
        super().__init__()
        self.setWindowTitle("Second Assignment")
        self.setFixedSize(310, 557)
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.center_on_screen()

        central = QWidget()
        central.setStyleSheet("""
            background-color: #4C4C4C;
            border-radius: 25px;
        """)
        self.setCentralWidget(central)

        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        # Store all input data here
        self.input_data = {
            "page1": {
                "start": 1,
                "end": 8,
                "probabilities": "0.125 0.125 0.125 0.125 0.125 0.125 0.125 0.125",
                "equal": True
            },
            "page2": {
                "start": 1,
                "end": 5,
                "probabilities": "0.2 0.2 0.2 0.2 0.2",
                "equal": True
            },
            "page3": {
                "instances": 20,
                "traffic_type": "Traffic"
            },
            "page4": {
                "output_option": "ð Excel"  # default selected option
            }
        }

        self.stacked = QStackedWidget()

        self.pages = [
            Page1(prev_callback=self.handle_back, next_callback=self.go_next, defaults=self.input_data["page1"]),
            Page2(prev_callback=self.go_prev, next_callback=self.go_next, defaults=self.input_data["page2"]),
            Page3(prev_callback=self.go_prev, next_callback=self.go_next, defaults=self.input_data["page3"]),
            Page4(prev_callback=self.go_prev, next_callback=self.go_next, defaults=self.input_data["page4"])
        ]

        for page in self.pages:
            self.stacked.addWidget(page)

        main_layout.addWidget(self.stacked)

        central.setLayout(main_layout)

    def handle_back(self):
        # If on first page, close window (return to main)
        if self.stacked.currentIndex() == 0:
            self.dashboard = DashboardWindow()
            self.dashboard.show()
            self.close()
        else:
            self.go_prev()

###########################################################################

class Page1(BasePage):
    def __init__(self, prev_callback=None, next_callback=None, defaults=None):
        super().__init__()

        defaults = defaults or {}

        # --- Title ---
        title_btn = QPushButton("Interarrival Time")
        title_btn.setEnabled(False)
        title_btn.setFixedSize(167, 30)
        title_btn.setStyleSheet("""
               background-color: #CDCDCD;
               border: 2px solid #BBBBBB;
               border-radius: 12px;
               font-size: 17px;
               color: #2C2C2C;
           """)
        self.main_layout.addLayout(self.centered_row(title_btn))

        # Fake space under title
        self.main_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Range section ---
        range_label = QLabel("Range")
        range_label.setStyleSheet("color: white; font-size: 14px;")

        self.start_field = self.make_field("Start")
        self.start_field.setText(str(defaults.get("start", "")))
        self.start_field.setValidator(QIntValidator(1, 10000))


        self.end_field = self.make_field("End")
        self.end_field.setText(str(defaults.get("end", "")))
        self.end_field.setValidator(QIntValidator(1, 10000))

        range_row = QHBoxLayout()
        range_row.addWidget(range_label)
        range_row.addStretch()
        range_row.addWidget(self.start_field)
        range_row.addWidget(self.end_field)
        self.main_layout.addLayout(range_row)

        self.main_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy  .Minimum, QSizePolicy.Fixed))

        # --- Equal section ---
        self.equal_checkbox = QCheckBox("Equal")
        self.equal_checkbox.setStyleSheet("color: white;")
        equal_text = QLabel("All probabilities are the same")
        equal_text.setStyleSheet("color: gray; font-size: 12px;")
        equal_text.setWordWrap(True)
        self.equal_checkbox.setChecked(defaults.get("equal", False))
        self.main_layout.addWidget(self.equal_checkbox)
        self.main_layout.addWidget(equal_text)
        self.main_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Probabilities section ---
        prob_label = QLabel("Probabilities")
        prob_label.setStyleSheet("color: white; font-size: 14px;")
        self.prob_field = self.make_field("Sum of Probs. must equal 1", width=200, height=40)
        self.prob_field.setText(defaults.get("probabilities", ""))


        regex = QRegularExpression(r"^(\d*\.?\d+\s*)+$")
        validator = QRegularExpressionValidator(regex)
        self.prob_field.setValidator(validator)
        self.main_layout.addWidget(prob_label)
        self.main_layout.addWidget(self.prob_field)

        # Push everything up
        self.main_layout.addStretch()

        # --- Nav buttons at bottom ---
        self.main_layout.addLayout(self.nav_layout)
        self.add_nav_buttons(prev_callback, next_callback)

###########################################################################

class Page2(BasePage):
    def __init__(self, prev_callback=None, next_callback=None, defaults=None):
        super().__init__()

        defaults = defaults or {}

        # --- Title ---
        title_btn = QPushButton("Service-Time")
        title_btn.setEnabled(False)
        title_btn.setFixedSize(167, 30)
        title_btn.setStyleSheet("""
               background-color: #CDCDCD;
               border: 2px solid #BBBBBB;
               border-radius: 12px;
               font-size: 17px;
               color: #2C2C2C;
           """)
        self.main_layout.addLayout(self.centered_row(title_btn))

        # Fake space under title
        self.main_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Range section ---
        range_label = QLabel("Range")
        range_label.setStyleSheet("color: white; font-size: 14px;")

        self.start_field = self.make_field("Start")
        self.start_field.setText(str(defaults.get("start", "")))
        self.start_field.setValidator(QIntValidator(1, 10000))

        self.end_field = self.make_field("End")
        self.end_field.setText(str(defaults.get("end", "")))
        self.end_field.setValidator(QIntValidator(1, 10000))

        range_row = QHBoxLayout()
        range_row.addWidget(range_label)
        range_row.addStretch()
        range_row.addWidget(self.start_field)
        range_row.addWidget(self.end_field)
        self.main_layout.addLayout(range_row)

        self.main_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Equal section ---
        self.equal_checkbox = QCheckBox("Equal")
        self.equal_checkbox.setStyleSheet("color: white;")
        equal_text = QLabel("All probabilities are the same")
        equal_text.setStyleSheet("color: gray; font-size: 12px;")
        equal_text.setWordWrap(True)
        self.equal_checkbox.setChecked(defaults.get("equal", False))
        self.main_layout.addWidget(self.equal_checkbox)
        self.main_layout.addWidget(equal_text)
        self.main_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Probabilities section ---
        prob_label = QLabel("Probabilities")
        prob_label.setStyleSheet("color: white; font-size: 14px;")
        self.prob_field = self.make_field("Sum of Probs. must equal 1", width=200, height=40)
        self.prob_field.setText(defaults.get("probabilities", ""))

        regex = QRegularExpression(r"^(\d*\.?\d+\s*)+$")
        validator = QRegularExpressionValidator(regex)
        self.prob_field.setValidator(validator)
        self.main_layout.addWidget(prob_label)
        self.main_layout.addWidget(self.prob_field)

        # Bottom stretch
        self.main_layout.addStretch()

        # --- Nav buttons at bottom ---
        self.main_layout.addLayout(self.nav_layout)
        self.add_nav_buttons(prev_callback, next_callback)

###########################################################################
class Page3(BasePage):
    def __init__(self, prev_callback=None, next_callback=None, defaults=None):
        super().__init__()
        defaults = defaults or {}

        # --- Title ---
        title_btn = QPushButton("Simulation Table")
        title_btn.setEnabled(False)
        title_btn.setFixedSize(167, 30)
        title_btn.setStyleSheet("""
               background-color: #CDCDCD;
               border: 2px solid #BBBBBB;
               border-radius: 12px;
               font-size: 17px;
               color: #2C2C2C;
           """)
        self.main_layout.addLayout(self.centered_row(title_btn))

        self.main_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- How Many Instances ---
        range_label = QLabel("How Many Instances")
        range_label.setStyleSheet("color: white; font-size: 14px;")

        self.num_instances_field = self.make_field("Ex.20")
        self.num_instances_field.setValidator(QIntValidator(1, 10000))
        self.num_instances_field.setText(str(defaults.get("instances", 20)))  # default = 20

        range_row = QHBoxLayout()
        range_row.addWidget(range_label)
        range_row.addStretch()
        range_row.addWidget(self.num_instances_field)
        self.main_layout.addLayout(range_row)

        self.main_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Radio section ---
        self.traffic_radio = QRadioButton("Traffic")
        self.traffic_radio.setChecked(defaults.get("traffic_selected", True))
        self.traffic_radio.setStyleSheet("color: white;")

        radio_layout = QHBoxLayout()
        radio_layout.addWidget(self.traffic_radio)
        radio_layout.addStretch()
        self.main_layout.addLayout(radio_layout)

        # Bottom stretch
        self.main_layout.addStretch()

        # --- Nav buttons ---
        self.main_layout.addLayout(self.nav_layout)
        self.add_nav_buttons(prev_callback, next_callback)

###########################################################################

class Page4(BasePage):
    def __init__(self, prev_callback=None, next_callback=None, defaults=None):
        super().__init__()
        defaults = defaults or {}

        # --- Title ---
        title_btn = QPushButton("Output")
        title_btn.setEnabled(False)
        title_btn.setFixedSize(167, 30)
        title_btn.setStyleSheet("""
            background-color: #CDCDCD;
            border: 2px solid #BBBBBB;
            border-radius: 12px;
            font-size: 17px;
            color: #2C2C2C;
        """)
        self.main_layout.addLayout(self.centered_row(title_btn))

        self.main_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # --- Radio Buttons with Read-only Descriptions ---
        options = [
            ("ð Excel", "Export results to an Excel (.xlsx) file"),
            ("ð¥ï¸ Python GUI", "Display results inside the app interface"),
            ("ð» Terminal", "Print results in the console output")
        ]

        self.button_group = QButtonGroup(self)
        self.button_group.setExclusive(True)
        self.radio_buttons = []

        default_selected = defaults.get("output_option", None)

        for label_text, description in options:
            radio = QRadioButton(label_text)
            radio.setStyleSheet("color: white; font-size: 14px;")
            self.button_group.addButton(radio)
            self.main_layout.addWidget(radio)

            # If this matches default, select it
            if label_text == default_selected:
                radio.setChecked(True)

            # --- Description field (read-only QLineEdit) ---
            desc_field = QLineEdit(description)
            desc_field.setReadOnly(True)
            desc_field.setFixedSize(300, 25)
            desc_field.setStyleSheet("""
                background-color: #3A3A3A;
                color: #E0E0E0;
                border: none;
                border-radius: 6px;
                padding-left: 6px;
                font-size: 13px;
            """)
            self.main_layout.addWidget(desc_field, alignment=Qt.AlignCenter)
            self.main_layout.addSpacerItem(QSpacerItem(20, 10, QSizePolicy.Minimum, QSizePolicy.Fixed))

        # Bottom stretch
        self.main_layout.addStretch()

        # --- Nav buttons at bottom ---
        self.main_layout.addLayout(self.nav_layout)
        self.add_nav_buttons(prev_callback, next_callback)


    def selected_option(self):
        """Return the text of the selected radio button, or None."""
        selected = self.button_group.checkedButton()
        return selected.text() if selected else None

    @staticmethod
    def generate_interarrival_distribution(start, end, probabilities=None):
        """
        Generate a 2D list for interarrival time distribution.

        Columns:
        [Interarrival Time, Probability, Cumulative Probability, Random-Digit Assignment]

        Random-digit assignment:
          starts at 001 and wraps so the final range ends at 000.
        """

        times = list(range(start, end + 1))
        n = len(times)

        # If equal probabilities selected
        if probabilities is None:
            probabilities = [1 / n] * n

        # Safety check
        if abs(sum(probabilities) - 1) > 1e-6:
            raise ValueError("Sum of probabilities must equal 1.")

        # Compute cumulative probabilities
        cumulative = []
        total = 0
        for p in probabilities:
            total += p
            cumulative.append(round(total, 3))

        # Assign random-digit ranges (001â000)
        random_ranges = []
        current_start = 1  # start from 001

        for i, p in enumerate(probabilities):
            span = round(p * 1000)
            if i == n - 1:
                # Last range ends at 000
                random_ranges.append(f"{current_start:03d} - 000")
            else:
                end_range = current_start + span - 1
                random_ranges.append(f"{current_start:03d} - {end_range:03d}")
                current_start = end_range + 1
                if current_start > 1000:
                    current_start -= 1000

        # Combine into 2D list
        table = []
        for i in range(n):
            row = [
                times[i],
                round(probabilities[i], 3),
                cumulative[i],
                random_ranges[i],
            ]
            table.append(row)

        # print("GID DEBUG TABLE:", table)
        return table

    @staticmethod
    def generate_service_time_distribution(start, end, probabilities=None):
        """
        Generate a 2D list for service-time distribution.

        Columns:
        [Service Time, Probability, Cumulative Probability, Random-Digit Assignment]

        Random-digit assignment:
          starts at 01 and wraps so the final range ends at 00 (1â100 scale).
        """

        times = list(range(start, end + 1))
        n = len(times)

        # If equal probabilities are not provided
        if probabilities is None:
            probabilities = [1 / n] * n

        # Validate
        if abs(sum(probabilities) - 1) > 1e-6:
            raise ValueError("Sum of probabilities must equal 1.")

        # Compute cumulative probabilities
        cumulative = []
        total = 0
        for p in probabilities:
            total += p
            cumulative.append(round(total, 3))

        # Assign random-digit ranges (01â00)
        random_ranges = []
        current_start = 1  # start from 01

        for i, p in enumerate(probabilities):
            span = round(p * 100)  # because 100 total digits
            if span == 0:
                span = 1  # ensure at least 1 digit

            if i == n - 1:
                random_ranges.append(f"{current_start:02d} - 00")
            else:
                end_range = current_start + span - 1
                random_ranges.append(f"{current_start:02d} - {end_range:02d}")
                current_start = end_range + 1
                if current_start > 100:
                    current_start -= 100

        # Combine into 2D table
        table = []
        for i in range(n):
            row = [
                times[i],
                round(probabilities[i], 3),
                cumulative[i],
                random_ranges[i],
            ]
            table.append(row)

        # print("GSTD Debug Table:", table)
        return table

    @staticmethod
    def assign_interarrival_times(distribution_table, num_users=10):
        """
        Generate a 2D list showing:
        [User, Random Digit, Interarrival Time]

        distribution_table: 2D list from generate_interarrival_distribution
        num_users: how many users to simulate
        """

        # --- Parse ranges from distribution_table ---
        parsed_ranges = []
        for time, prob, cum_prob, r_range in distribution_table:
            start_str, end_str = r_range.replace(' ', '').split('-')
            start = int(start_str)
            end = 1000 if end_str == '000' else int(end_str)
            parsed_ranges.append((time, start, end))

        # --- Generate random digits and assign times ---
        table = []
        for user_id in range(1, num_users + 1):
            random_digit = random.randint(1, 1000)

            # Find which interval this random digit belongs to
            interarrival_time = None
            for time, start, end in parsed_ranges:
                if start <= random_digit <= end or (end == 1000 and random_digit == 1000):
                    interarrival_time = time
                    break
                # wrap case (001â000)
                if start > end and (random_digit >= start or random_digit <= end):
                    interarrival_time = time
                    break

            table.append([user_id, f"{random_digit:03d}", interarrival_time])

        # print("AIT Debug Table:", table)
        return table

    @staticmethod
    def assign_service_times(distribution_table, num_customers=10):
        """
        Generate a 2D list:
        [Customer, Random Digit, Service Time]

        distribution_table: from generate_service_distribution
        num_customers: number of customers to simulate
        """

        parsed_ranges = []
        for time, prob, cum_prob, r_range in distribution_table:
            start_str, end_str = r_range.replace(' ', '').split('-')
            start = int(start_str)
            end = 100 if end_str == '00' else int(end_str)
            parsed_ranges.append((time, start, end))

        # --- Generate random digits and assign service times ---
        table = []
        for cust_id in range(1, num_customers + 1):
            random_digit = random.randint(1, 100)

            service_time = None
            for time, start, end in parsed_ranges:
                if start <= random_digit <= end or (end == 100 and random_digit == 100):
                    service_time = time
                    break
                # wrap around case like 84â00
                if start > end and (random_digit >= start or random_digit <= end):
                    service_time = time
                    break

            table.append([cust_id, f"{random_digit:02d}", service_time])

        # print("AST Debug Table:", table)
        return table

    @staticmethod
    def simulate_queue(interarrival_dist, service_dist, num_users):
        inter_table = Page4.assign_interarrival_times(interarrival_dist, num_users)
        service_table = Page4.assign_service_times(service_dist, num_users)

        simulation = []
        server_available_time = 0

        for i in range(num_users):
            user = i + 1
            interarrival_time = float(inter_table[i][2])
            service_time = float(service_table[i][2])

            if i == 0:
                arrival_time = interarrival_time
            else:
                arrival_time = simulation[i - 1][2] + interarrival_time

            service_begin = max(arrival_time, server_available_time)
            waiting_time = service_begin - arrival_time
            service_end = service_begin + service_time
            time_in_system = waiting_time + service_time
            idle_time = max(0, arrival_time - server_available_time)

            server_available_time = service_end

            simulation.append([
                user,
                interarrival_time,
                arrival_time,
                service_time,
                service_begin,
                waiting_time,
                service_end,
                time_in_system,
                idle_time
            ])

        # --- Compute Performance Metrics ---
        total_waiting = sum(row[5] for row in simulation)
        total_service = sum(row[3] for row in simulation)
        total_time_in_system = sum(row[7] for row in simulation)
        total_idle = sum(row[8] for row in simulation)

        num_waited = sum(1 for row in simulation if row[5] > 0)
        total_customers = len(simulation)

        total_simulation_time = simulation[-1][6]  # time when last service ended

        avg_waiting_time = total_waiting / total_customers
        prob_waiting = num_waited / total_customers
        avg_service_time = total_service / total_customers
        avg_time_in_system = total_time_in_system / total_customers
        server_utilization = total_service / total_simulation_time
        prob_server_idle = 1 - server_utilization

        metrics = {
            "Average Waiting Time": round(avg_waiting_time, 3),
            "Probability of Waiting": round(prob_waiting, 3),
            "Average Service Time": round(avg_service_time, 3),
            "Average Time in System": round(avg_time_in_system, 3),
            "Server Utilization": round(server_utilization, 3),
            "Probability Server Idle": round(prob_server_idle, 3)
        }

        # print("SQ Debug Table:", simulation)
        Page4.print_table_terminal(
            inter_table, ["User", "Random", "Interarrival Time"]
        )
        Page4.print_table_terminal(
            service_table, ["User", "Random", "Service Time"]
        )
        print("Queue Metrics:", metrics)

        return [simulation, metrics]

    # Function 1: Print table in terminal
    @staticmethod
    def print_table_terminal(table, headers=None):
            """
            Prints a neatly aligned 2D table in the terminal with optional headers.
            Automatically adjusts column widths and handles numeric values.
            """
            if not table:
                print("(Empty table)")
                return

            # Convert all cells to strings
            str_table = [[str(cell) for cell in row] for row in table]

            # Include headers if provided
            if headers:
                str_headers = [str(h) for h in headers]
                data = [str_headers] + str_table
            else:
                data = str_table

            # Compute column widths
            col_widths = [max(len(row[i]) for row in data) for i in range(len(data[0]))]

            # Helper to format a row
            def format_row(row):
                return " | ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(row)))

            # Print header
            if headers:
                print(format_row(headers))
                print("-" * (sum(col_widths) + 3 * (len(col_widths) - 1)))

            # Print each row
            for row in table:
                print(format_row(row))

    # Function 2: Save table to .txt file
    def save_table_txt(table, filename="table_output.txt", headers=None):
        """
        Saves a 2D table to a text file.
        """
        with open(filename, "w") as f:
            if headers:
                f.write(" | ".join(map(str, headers)) + "\n")
                f.write("-" * (len(headers) * 10) + "\n")
            for row in table:
                f.write(" | ".join(map(str, row)) + "\n")
        print(f"Table saved to {filename}")

    # Function 3: Save table to Excel (.xlsx)
    def save_table_excel(table, filename="table_output.xlsx", headers=None):
        """
        Saves a 2D table to Excel using openpyxl.
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("Please install openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active

        row_offset = 1
        if headers:
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            row_offset = 2

        for r, row in enumerate(table, start=row_offset):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # Optional: auto-fit column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        wb.save(filename)
        print(f"Table saved to {filename}")

